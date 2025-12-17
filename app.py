import streamlit as st
import pandas as pd
import io
import openpyxl
import os
from datetime import datetime
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill, Protection, numbers
import numpy as np
import re
import base64

# --- FUNCIONES AUXILIARES ---

def clean_leading_zeros(code_str):
    if pd.isna(code_str) or code_str is None:
        return ''
    code_str = str(code_str).strip()
    if not code_str:
        return ''
    return code_str.lstrip('0')

def clean_input_codes(input_raw):
    if not input_raw:
        return []
    codes = [c.strip() for c in input_raw.split(',') if c.strip()]
    cleaned_codes = [clean_leading_zeros(c) for c in codes]
    return list(set([c for c in cleaned_codes if c]))

def detect_portfolio_code(df):
    # ESTRATEGIA 1: Buscar por columna "Organización de Ventas"
    col_org_venta = None
    for col in df.columns:
        c_clean = str(col).lower().replace(' ', '').replace('.', '').replace('_', '')
        if 'org' in c_clean and ('ven' in c_clean or 'vta' in c_clean):
            col_org_venta = col
            break
            
    if col_org_venta:
        unique_vals = set(df[col_org_venta].astype(str).str.strip().str.upper().unique())
        unique_set = set()
        for val in unique_vals:
            unique_set.add(val) 
            unique_set.add(clean_leading_zeros(val)) 
        
        if '0702' in unique_set or '702' in unique_set: return '0700'
        if '0602' in unique_set or '602' in unique_set: return '0600'
        if 'R200' in unique_set: return 'R100'
        if 'C001' in unique_set: return 'C001'

    # ESTRATEGIA 2: Buscar por columna "Sociedad"
    col_sociedad = None
    for col in df.columns:
        c_clean = str(col).lower().replace(' ', '')
        if 'sociedad' in c_clean:
            col_sociedad = col
            break
    
    if col_sociedad:
        unique_soc = set(df[col_sociedad].astype(str).str.strip().str.lower().unique())
        for val in unique_soc:
            if 'alimentos' in val and 'polar' in val: return '0700'
            if 'pepsi' in val: return 'R100'
            if 'cervecer' in val or 'cerveceria' in val: return 'C001'
            if 'efe' in val: return '0600'

    # ESTRATEGIA 3: Fallback por Clase de Factura
    clase_factura_keys = ['clase de factura', 'clasefactura', 'clase_factura', 'clase.factura', 'cl.f'] 
    col_clase_factura = next((c for c in df.columns if any(k in str(c).lower().replace(' ', '') for k in clase_factura_keys)), None)

    if col_clase_factura:
        unique_factura_codes = set(df[col_clase_factura].astype(str).str.strip().str.upper().unique())
        if any(code in unique_factura_codes for code in ['YP01', 'YP04', 'YP10']): return 'R100'
        if 'YC00' in unique_factura_codes: return 'C001'
        if any(code in unique_factura_codes for code in ['ZSPN', 'X|', 'ZSCC']): return '0700'

    return '--'

def convert_value_to_float(value):
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        processed_value = value.strip().replace('$', '').replace('€', '').replace(' ', '')
        if ',' in processed_value and '.' in processed_value:
            processed_value = processed_value.replace('.', '')
            processed_value = processed_value.replace(',', '.')
        elif ',' in processed_value:
            processed_value = processed_value.replace(',', '.')
        processed_value = re.sub(r'[^\d.]', '', processed_value)
        try:
            if processed_value.startswith('.'): 
                processed_value = '0' + processed_value
            if processed_value.endswith('.'): 
                processed_value = processed_value[:-1]
            return float(processed_value)
        except ValueError:
            return None
    return None

def format_monto_local(monto):
    if pd.isna(monto) or monto is None:
        return ''
    try:
        monto = float(monto)
        return f"{monto:,.2f}".replace(",", "_TEMP_").replace(".", ",").replace("_TEMP_", ".")
    except (ValueError, TypeError):
        return str(monto) 

def find_invoices_by_total_sum(df_candidates, target_amount, invoice_col, price_col, client_col, product_col, assignment_mode):
    if df_candidates.empty or invoice_col not in df_candidates.columns or price_col not in df_candidates.columns:
        return None, None, 0

    df_candidates_copy = df_candidates.copy() 
    if '__monto_numeric__' not in df_candidates_copy.columns:
        df_candidates_copy['__monto_numeric__'] = df_candidates_copy[price_col].apply(convert_value_to_float)
    
    df_candidates_copy.dropna(subset=['__monto_numeric__'], inplace=True)
    df_candidates_copy = df_candidates_copy[df_candidates_copy['__monto_numeric__'] > 0.01]

    if df_candidates_copy.empty:
        return None, None, 0

    invoice_sums_df = df_candidates_copy.groupby(invoice_col).agg(
        total_sum=('__monto_numeric__', 'sum'),
        client_code=(client_col, 'first'),
        product_code=(product_col, 'first') 
    ).reset_index()

    invoice_sums_df.columns = [invoice_col, 'total_sum', client_col, product_col]

    sufficient_invoices = invoice_sums_df[invoice_sums_df['total_sum'] >= target_amount]

    if not sufficient_invoices.empty:
        best_single_invoice_df = sufficient_invoices.sort_values(by='total_sum', ascending=True).iloc[0:1]
        df_selected_invoices = pd.DataFrame(best_single_invoice_df)
    else:
        invoice_sums_df.sort_values(by='total_sum', ascending=False, inplace=True)
        chosen_invoices_data = []
        current_sum = 0
        for _, row in invoice_sums_df.iterrows():
            if current_sum >= target_amount:
                break
            chosen_invoices_data.append(row.to_dict())
            current_sum += row['total_sum']
        df_selected_invoices = pd.DataFrame(chosen_invoices_data)
    
    if df_selected_invoices.empty:
        total_available_sum = invoice_sums_df['total_sum'].sum()
        return None, None, total_available_sum

    current_sum = df_selected_invoices['total_sum'].sum()
    if current_sum < target_amount:
        total_available_sum = invoice_sums_df['total_sum'].sum()
        return None, None, total_available_sum

    first_client_code = df_selected_invoices[client_col].iloc[0]
    monto_cubierto_final = df_selected_invoices['total_sum'].sum()

    if assignment_mode == 'Estricto (Truncar)' and not df_selected_invoices.empty:
        sum_excluding_last = df_selected_invoices['total_sum'].iloc[:-1].sum() if len(df_selected_invoices) > 1 else 0
        required_from_last = target_amount - sum_excluding_last
        required_from_last = max(0, required_from_last)

        df_selected_invoices_copy = df_selected_invoices.copy()
        last_idx_in_df = df_selected_invoices_copy.index[-1]
        df_selected_invoices_copy.loc[last_idx_in_df, 'total_sum'] = required_from_last
        
        monto_cubierto_final = target_amount
        df_selected_invoices_copy['Monto NC Asignado'] = df_selected_invoices_copy['total_sum']
        return df_selected_invoices_copy, first_client_code, monto_cubierto_final
    else:
        df_selected_invoices['Monto NC Asignado'] = df_selected_invoices['total_sum']
        return df_selected_invoices, first_client_code, monto_cubierto_final

def create_excel_for_all_invoices(df_to_export, selected_portfolio, ticket_number_for_name="", multiple_invoices=False):
    portfolio_template_map = {
        '0700': 'plantilla_APC.xlsx',
        'R100': 'plantilla_PCV.xlsx',
        'C001': 'plantilla_CYM.xlsx',
        '0600': 'plantilla_EFE.xlsx',
    }
    
    try:
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        BASE_DIR = os.getcwd()

    default_template_path = os.path.join(BASE_DIR, 'plantilla_default.xlsx')
    specific_template_name = portfolio_template_map.get(selected_portfolio)
 
    template_path = os.path.join(BASE_DIR, specific_template_name) if specific_template_name and os.path.exists(os.path.join(BASE_DIR, specific_template_name)) else default_template_path
    
    try:
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook.active
    except Exception as e:
        try:
             workbook = openpyxl.load_workbook(default_template_path)
             sheet = workbook.active
        except Exception as e_default:
             st.error(f"Error: No se pudo cargar ninguna plantilla de Excel (ni específica ni por defecto). Detalles: {e_default}")
             return None
             
    header_row = 1
    df_column_map = {
        "Clase de pedido": "Clase de pedido",
        "Organizacion de Venta": "Organizacion de Venta",
        "Canal de Distribucion": "Canal de Distribucion",
        "Sector": "Sector",
        "Solicitante": "Solicitante",
        "Fecha de Pedido": "Fecha de Pedido",
        "Fecha de Precio": "Fecha de Precio",
        "Fecha de Factura": "Fecha de Factura",
        "Motivo": "Motivo",
        "Pedido Cliente": "Pedido Cliente",
        "Material": "Material",
        "Cantidad": "Cantidad",
        "U. MEDIDA": "U. MEDIDA", 
        "CONDICION": "CONDICION", 
        "VARIACION DE PRECIO": "VARIACION DE PRECIO",
        "ASIGNACION": "ASIGNACION",
        "UTILIZACION": "UTILIZACION",
        "TEXTO CABECERA": "TEXTO CABECERA",
        "Peso %": "Peso %",
        "Monto NC Asignado": "Monto NC Asignado",
        "Observación": "Observación",
    }
    excel_template_structure = []
    template_headers = {}
    used_df_columns = set()

    for cell in sheet[header_row]:
        template_header = str(cell.value).strip() if cell.value is not None else ''
        if not template_header: 
            continue
        normalized_template_header = template_header.replace(' ', '').lower()
        df_column_name = None
        
        for df_col, template_col_match in df_column_map.items():
            if template_col_match.replace(' ', '').lower() == normalized_template_header:
                df_column_name = df_col
                break
        
        if df_column_name is None:
            if "clase" in normalized_template_header and "pedido" in normalized_template_header:
                df_column_name = "Clase de pedido"
            elif "org" in normalized_template_header and ("ven" in normalized_template_header or "vta" in normalized_template_header):
                df_column_name = "Organizacion de Venta"
            elif "canal" in normalized_template_header:
                df_column_name = "Canal de Distribucion"
            elif "sector" in normalized_template_header:
                df_column_name = "Sector"

        if df_column_name == "TEXTO CABECERA" and "TEXTO CABECERA" in used_df_columns:
            continue

        excel_template_structure.append({
            'letter': cell.column_letter,
            'df_column': df_column_name,
        })
        template_headers[template_header] = cell.column_letter
        
        if df_column_name:
            used_df_columns.add(df_column_name)
    
    start_row = 2
    max_rows = sheet.max_row
    
    if max_rows >= start_row:
        sheet.delete_rows(start_row, max_rows - start_row + 1)
        
    if df_to_export.empty:
        output_buffer = io.BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer

    required_df_cols = [col for col in df_column_map.keys() if col is not None]
    df_final = df_to_export.reindex(columns=required_df_cols, fill_value='')

    NUMBER_FORMAT = '0.00' 
    
    for row_idx, data_tuple in enumerate(df_final.itertuples(index=False), start=start_row):
        for col_data in excel_template_structure:
            excel_col_letter = col_data['letter']
            df_col_name = col_data['df_column']
            cell_value = None 

            if df_col_name and df_col_name in df_final.columns:
                try:
                    col_position_in_tuple = df_final.columns.get_loc(df_col_name)
                    value = data_tuple[col_position_in_tuple]
                    
                    if pd.isna(value) or value == '':
                        cell_value = ''
                    elif df_col_name in ["Fecha de Pedido", "Fecha de Precio", "Fecha de Factura"]:
                        cell_value = str(value) 
                    elif df_col_name in ["VARIACION DE PRECIO", "Monto NC Asignado"]:
                        try:
                            numeric_value = convert_value_to_float(value)
                            if numeric_value is not None:
                                cell_value = numeric_value
                                sheet[f"{excel_col_letter}{row_idx}"].number_format = NUMBER_FORMAT 
                            else:
                                cell_value = str(value)
                        except Exception:
                             cell_value = str(value)
                    else:
                        cell_value = str(value)
                except KeyError:
                     cell_value = ''
                except IndexError as ie:
                     st.error(f"Error de índice al acceder a la tupla para la columna '{df_col_name}' en la fila {row_idx}: {ie}. Data tuple length: {len(data_tuple)}, requested index: {col_position_in_tuple}")
                     cell_value = ''

            if cell_value is not None:
                sheet[f"{excel_col_letter}{row_idx}"] = cell_value
            
            if df_col_name in ["ASIGNACION", "TEXTO CABECERA", "Observación", "Solicitante", "Material", "Pedido Cliente"]:
                 sheet[f"{excel_col_letter}{row_idx}"].alignment = Alignment(horizontal='left')
            if df_col_name in ["Cantidad", "Peso %", "VARIACION DE PRECIO", "Monto NC Asignado"]:
                sheet[f"{excel_col_letter}{row_idx}"].alignment = Alignment(horizontal='right')

    cols_to_resize = [
        "Clase de pedido", "Organizacion de Venta", "Canal de Distribucion", "Sector", "Solicitante", 
        "Fecha de Pedido", "Fecha de Precio", "Fecha de Factura", "Material", "Pedido Cliente",
        "Cantidad", "U. MEDIDA", "CONDICION", "VARIACION DE PRECIO", "ASIGNACION", "Peso %", 
        "Monto NC Asignado"
    ]
    large_text_cols = {
        "VARIACION DE PRECIO": 18,
        "TEXTO CABECERA": 35,
        "Pedido Cliente": 35,
        "Observación": 25,
        "Motivo": 8 
    }
    default_width = 15
    
    for col_name, col_letter in template_headers.items():
        if col_name in large_text_cols:
             sheet.column_dimensions[col_letter].width = large_text_cols[col_name]
        elif col_name in cols_to_resize:
            sheet.column_dimensions[col_letter].width = default_width
    
    output_buffer = io.BytesIO()
    
    try:
        workbook.save(output_buffer)
        output_buffer.seek(0)
    except Exception as e:
        st.error(f"Error al guardar el archivo Excel: {e}")
        return None
        
    return output_buffer

@st.cache_data
def load_simple_table(uploaded_file):
    try:
        df = pd.read_excel(uploaded_file, dtype=str)
    except Exception:
        try:
            uploaded_file.seek(0)
            df = pd.read_csv(uploaded_file, sep=',', dtype=str)
        except Exception:
            try:
                 uploaded_file.seek(0)
                 df = pd.read_csv(uploaded_file, sep=';', dtype=str)
            except Exception as e:
                 st.error(f"Error al cargar el archivo. Asegúrese de que sea un Excel .xlsx, .xls o CSV: {e}")
                 return None
                 
    df.columns = df.columns.str.strip()
    df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    return df

def get_file_name(portfolio_cod, ticket, is_first=False, multiple_invoices=False):
    global BASE_DIR
    try:
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    except NameError:
        BASE_DIR = os.getcwd()

    PORTFOLIO_ACRONYM_MAP = {
        '0700': 'APC',
        'R100': 'PCV',
        'C001': 'CYM',
        '0600': 'EFE',
        '--': 'SIN_PORTAFOLIO'
    }

    acronym = PORTFOLIO_ACRONYM_MAP.get(portfolio_cod, 'SIN_PORTAFOLIO')
    if multiple_invoices:
        return f"NC_Multiples_{acronym}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    elif ticket:
        return f"TICKET#SR-{ticket}-{acronym}.xlsx"
    else:
        return f"TICKET_SIN_NUMERO-{acronym}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# --- FUNCIÓN DE LIMPIEZA ---
def clear_form_data():
    keys_text = [
        'filtro_cliente_cod',
        'filtro_producto_cod',
        'widget_monto_nc', 
        'filtro_ticket_cod',
        'monto_display'
    ]
    for key in keys_text:
        if key in st.session_state:
            st.session_state[key] = ""
            
    if 'filtro_motivo' in st.session_state:
        del st.session_state['filtro_motivo']
        
    if 'filtro_monto' in st.session_state:
        st.session_state['filtro_monto'] = None

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(
    page_title="Notas de Crédito", 
    page_icon="🔴🔵", 
    layout="wide",
    initial_sidebar_state="expanded", # Sidebar abierta por defecto
    menu_items={
        'Get Help': 'https://www.google.com',
        'Report a bug': 'https://www.google.com',
        'About': 'Aplicación de Notas de Crédito'
    }
) 

st.markdown("""
    <style>
        header { visibility: hidden; }
        [data-testid="stDecoration"] { visibility: hidden; }
        [data-testid="stHeader"] { background-color: transparent; }
        
        .block-container { padding-top: 1rem !important; }

        section[data-testid="stSidebar"] [data-testid="stFileUploader"],
        section[data-testid="stSidebar"] [data-testid="stButton"],
        section[data-testid="stSidebar"] [data-testid="stTextInput"],
        section[data-testid="stSidebar"] [data-testid="stSelectbox"] {
            margin-bottom: 12px !important;
        }

        [data-testid="stSidebar"] [data-testid="stFileUploader"] section[data-testid="stFileUploadDropzone"] {
            min-height: auto !important;
            padding: 0.75rem !important; 
        }

        [data-testid="stSidebar"] [data-testid="stFileUploader"] [data-testid="stFileUploaderInstructions"] {
            padding: 0 !important;
        }

        [data-testid="stSidebar"] [data-testid="stFileUploader"] p {
            font-size: 0.8rem;
            margin-bottom: 0.25rem;
        }

        [data-testid="stSidebar"] [data-testid="stFileUploader"] button {
            padding: 2px 8px;
            font-size: 0.8rem;
        }

        section[data-testid="stSidebar"] {
            padding-top: 5px !important;
        }
        
        section[data-testid="stSidebar"] button {
            background-color: white !important;
            border-color: #00449C !important;
            color: #00449C !important;
        }
        section[data-testid="stSidebar"] button:hover {
            background-color: #f0f2f6 !important;
            border-color: #00449C !important;
            color: #00449C !important;
        }
        
        section[data-testid="stSidebar"] [data-testid="stFileUploader"] { margin-bottom: 12px !important; }
        section[data-testid="stSidebar"] [data-testid="stButton"] { margin-bottom: 12px !important; }
        [data-testid="stForm"] { margin-top: 0px !important; border: none !important; padding: 0 !important; }
        
        [data-testid="stSidebar"] .stMarkdown, [data-testid="stSidebar"] label { color: initial !important; }
        [data-testid="stSidebar"] .stRadio label span { color: initial; }

        [data-testid="stSidebar"] div.stSelectbox, [data-testid="stSidebar"] div.stTextInput {
            background-color: white !important;
            border-radius: 5px;
        }
        [data-testid="stSidebar"] div.stSelectbox div[data-testid="stInputContainer"],
        [data-testid="stSidebar"] div.stTextInput div[data-testid="stInputContainer"] {
            background-color: white !important;
        }

        [data-testid="stSidebar"] [data-testid="stImage"] {
            display: flex; justify-content.center; margin-left: auto; margin-right: auto; margin-top: 5px;
        }
        
        [data-testid="stSidebar"] [data-testid="stFileUploader"] ul { display: none; }
        [data-testid="stSidebar"] [data-testid="stFileUploader"] .st-emotion-cache-1pxpzwy,
        [data-testid="stSidebar"] [data-testid="stFileUploader"] .st-emotion-cache-1pxpzwy p { color: #00449C; }
        [data-testid="stSidebar"] [data-testid="stFileUploader"] .st-emotion-cache-1pxpzwy svg { fill: #00449C; }
        
        .logo-container-right { display: none; }
        [data-testid="stHorizontalBlock"] h1 { margin-top: 0px !important; padding-top: 0px !important; }

        li[role="option"] span {
            white-space: normal !important; line-height: 1.2 !important; height: auto !important;
            overflow: visible !important; text-overflow: clip !important;
        }
        
        div[data-baseweb="popover"] { max-width: 90vw !important; }
    </style>
""", unsafe_allow_html=True)

try:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
except NameError:
    BASE_DIR = os.getcwd()

PORTFOLIO_ACRONYM_MAP = {
    '0700': 'APC',
    'R100': 'PCV',
    'C001': 'CYM',
    '0600': 'EFE',
    '--': 'SIN_PORTAFOLIO'
}

ALLOWED_INVOICE_CLASSES = {
    '0700': ['ZSPN', 'X|', 'ZSCC'],
    'R100': ['YP01', 'YP04', 'YP10'],
    'C001': ['YC00'],
    '0600': ['ZSPN', 'ZSCC'], 
}

LOGO_FILENAME_MAP = {
    '0700': 'Alimentos Polar (Completo).webp',
    'R100': 'Pepsi-Cola.webp',
    'C001': 'Cervecería Polar (Completo).webp',
    '0600': 'Productos EFE.webp',
    '--': 'image_8085bf.png', 
    '0700_SOLO': 'Alimentos Polar (Solo Logo).webp',
    'R100_SOLO': 'Pepsi-Cola (Solo Logo).webp', 
    'C001_SOLO': 'Cervecería Polar (Solo Logo).webp',
    '0600_SOLO': 'Productos EFE (Solo Logo).webp', 
    'EMPRESAS_POLAR': 'image_8085bf.png' 
}


if 'df_full' not in st.session_state:
    st.session_state.df_full = None
if 'monto_display' not in st.session_state:
    st.session_state['monto_display'] = ''
if 'assignment_mode' not in st.session_state:
    st.session_state['assignment_mode'] = 'Prorrateo (Recomendado)'
if 'filtro_producto_cod' not in st.session_state:
    st.session_state['filtro_producto_cod'] = ''
if 'portafolio_cod' not in st.session_state:
    st.session_state['portafolio_cod'] = '--' 
if 'stacked_invoices' not in st.session_state:
    st.session_state['stacked_invoices'] = []


limpiar_button = False

with st.sidebar:
    local_limpiar_button = False 
    
    uploaded_file = st.file_uploader(
        "Cargar archivo", 
        type=['xlsx', 'xls', 'csv'], 
        label_visibility="collapsed"
    ) 

    if uploaded_file and st.session_state.get('df_full') is None:
        full_name = uploaded_file.name
        short_name = full_name[:25] + "..." if len(full_name) > 25 else full_name
        st.info(f"📄 Archivo: {short_name}")
    
    analyze_button = st.button("Analizar y Cargar", disabled=(uploaded_file is None), use_container_width=True, type="primary")

    if analyze_button and uploaded_file:
        with st.spinner("Procesando archivo..."):
            df_loaded = load_simple_table(uploaded_file)
            st.session_state.df_full = df_loaded
            st.session_state.file_name = uploaded_file.name
            
            if df_loaded is not None:
                detected_code = detect_portfolio_code(df_loaded.copy())
                st.session_state['portafolio_cod'] = detected_code
            
        st.rerun()

    if st.session_state.get('df_full') is not None:
        
        with st.form(key='parametros_nc_form'):
            motivo_options = [
                "Anulación documento",
                "Avisos.Cabezales.Publicidad",
                "Descuento no Reflejado",
                "Diferencia en Precio",
                "Grand Slam",
                "Promoción.Sell Out.Reconocimiento",
                "Sin Motivo"
            ]
            st.selectbox("Motivo:", options=motivo_options, key="filtro_motivo")
            
            st.text_input("Cod. Cliente:", key="filtro_cliente_cod")
            
            st.text_input("Cod. Producto:", key="filtro_producto_cod")
          
            monto_input_str = st.text_input(
                "Monto NC:", 
                value=st.session_state.get('monto_display', ''),
                placeholder="",
                key="widget_monto_nc"
            )
            monto_nc_temp = convert_value_to_float(monto_input_str)
            st.session_state['filtro_monto'] = monto_nc_temp
            
            if monto_nc_temp is not None:
                try:
                     st.session_state['monto_display'] = format_monto_local(monto_nc_temp)
                except ValueError:
                    st.session_state['monto_display'] = monto_input_str
            else:
                st.session_state['monto_display'] = monto_input_str
                
            st.text_input("N° de Ticket:", key="filtro_ticket_cod")
            
            st.session_state['assignment_mode'] = 'Prorrateo (Recomendado)'
            
            submit_button = st.form_submit_button("Cargar", use_container_width=True)

        local_limpiar_button = st.button("Limpiar", use_container_width=True, on_click=clear_form_data)
        
    limpiar_button = local_limpiar_button

col_config_dict = {
    "Cod. Cliente": st.column_config.TextColumn("Cod. Cliente", width="small", help="Código del Cliente (Solicitante)"),
    "Factura": st.column_config.TextColumn("Factura", width="small", help="Número de Factura o Documento de Asignación"),
    "Cod. Producto": st.column_config.TextColumn("Cod. Producto", width="small", help="Código del Material o Producto"),
    "U. Medida": st.column_config.TextColumn("U. Medida", width="small", help="Unidad de Medida (Ej: UN, KG)"),
    "Monto Total Factura": st.column_config.TextColumn("Monto Total Factura", width="small", help="Suma de las líneas de la factura original."),
    "Monto Nota de Crédito": st.column_config.TextColumn("Monto Nota de Crédito", width="small", help="Monto de la NC asignado (prorrateado) a esta factura."),
}

tab1, tab2 = st.tabs(["Generador de Notas de Crédito", "Tickets Generados"])

with tab1:
    if st.session_state.get('df_full') is None:
        empresas_polar_logo_filename = LOGO_FILENAME_MAP.get('EMPRESAS_POLAR')
        logo_base64 = None
        if empresas_polar_logo_filename and os.path.exists(os.path.join(BASE_DIR, empresas_polar_logo_filename)):
            with open(os.path.join(BASE_DIR, empresas_polar_logo_filename), "rb") as f:
                data = f.read()
                logo_base64 = base64.b64encode(data).decode()

        if logo_base64:
            st.markdown(f"""
                <div style="display: flex; align-items: center; gap: 20px; margin-bottom: 2rem;">
                    <h1 style="margin: 0;">Notas de Crédito</h1>
                    <img src="data:image/png;base64,{logo_base64}" width="150">
                </div>
                """, unsafe_allow_html=True)
        else:
            st.title("Notas de Crédito")
        
        st.info("Por favor, cargue un archivo para comenzar.")
    else:
        current_portfolio_cod_display = st.session_state.get('portafolio_cod', '--')
        logo_key_solo = f"{current_portfolio_cod_display}_SOLO"
        logo_filename_dynamic = LOGO_FILENAME_MAP.get(logo_key_solo)
        if not logo_filename_dynamic or not os.path.exists(os.path.join(BASE_DIR, logo_filename_dynamic)):
            logo_key_full = current_portfolio_cod_display
            logo_filename_dynamic = LOGO_FILENAME_MAP.get(logo_key_full)

        logo_base64 = None
        if logo_filename_dynamic and os.path.exists(os.path.join(BASE_DIR, logo_filename_dynamic)):
            with open(os.path.join(BASE_DIR, logo_filename_dynamic), "rb") as f:
                data = f.read()
                logo_base64 = base64.b64encode(data).decode()

        if logo_base64:
            st.markdown(f"""
                <div style="display: flex; align-items: center; gap: 20px; margin-bottom: 2rem;">
                    <h1 style="margin: 0;">Notas de Crédito</h1>
                    <img src="data:image/webp;base64,{logo_base64}" width="100">
                </div>
                """, unsafe_allow_html=True)
        else:
            st.title("Notas de Crédito")
        
        df = st.session_state.df_full.copy()
        df_para_mostrar = pd.DataFrame()
        monto_nc = st.session_state.get('filtro_monto')
        
        condicion_forzada = None  
        template_condicion = 'ZNOT'

        PORTFOLIO_DEFAULTS = {
            'R100': {
                'Clase de pedido': 'YNCR', 
                'Organizacion de Venta': 'R200', 
                'Canal de Distribucion': 'FB', 
                'Sector': 'E2',                
                'CONDICION': 'YLIQ'            
            },
            'C001': {
                'Clase de pedido': 'ZCDF', 
                'Organizacion de Venta': 'C001', 
                'Canal de Distribucion': 'FB', 
                'Sector': 'E2',                
                'CONDICION': 'ZXAM'            
            },
            '0700': {
                'Clase de pedido': 'Z1MA', 
                'Organizacion de Venta': '0702', 
                'Canal de Distribucion': 'FB', 
                'Sector': 'E2',
                'CONDICION': 'ZNOT'
            },
            '0600': {
                'Clase de pedido': 'Z1MA', 
                'Organizacion de Venta': '0602', 
                'Canal de Distribucion': 'FB', 
                'Sector': 'E2',                
                'CONDICION': 'ZNOT'
            },
        }

        selected_portfolio_cod = st.session_state.get('portafolio_cod')
        assignment_mode = st.session_state.get('assignment_mode') 
        ticket_number = st.session_state.get('filtro_ticket_cod', '').strip() 

        client_code_input_raw = st.session_state.get('filtro_cliente_cod', '').strip()
        client_code_list = clean_input_codes(client_code_input_raw) 
        
        product_code_input_raw = st.session_state.get('filtro_producto_cod', '').strip()
        product_code_list = clean_input_codes(product_code_input_raw) 

        cliente_a_usar = None
        
        st.session_state['df_for_export_single_line'] = None
        df_para_mostrar_editor = pd.DataFrame() 

        try:
            df_pre_filtros = df.copy()
            
            factura_keys = ['factura', 'nofactura', 'numerofactura', 'asignacion']
            col_factura = next((c for c in df_pre_filtros.columns if any(k in str(c).lower().replace(" ", "").replace("°", "") for k in factura_keys)), None)

            # --- DETECCIÓN DE COLUMNAS ANTES DE LA LÓGICA DE SALDO ---
            col_monto = next((c for c in df_pre_filtros.columns if 'precio' in str(c).lower() and 'total' not in str(c).lower()), None)
            if col_monto is None:
                 col_monto = next((c for c in df_pre_filtros.columns if any(k in str(c).lower() for k in ['precio', 'neto', 'valor', 'monto']) and 'total' not in str(c).lower()), None)
            if col_monto is None:
                 col_monto = next((c for c in df_pre_filtros.columns if any(k in str(c).lower() for k in ['total'])), None)
            
            cliente_keys = ['cliente', 'codcliente', 'solicitante']
            col_cliente = next((c for c in df_pre_filtros.columns if any(k in str(c).lower() for k in cliente_keys)), None)
            
            producto_keys = ['producto', 'material', 'codigoproducto'] 
            col_producto = next((c for c in df_pre_filtros.columns if any(k in str(c).lower() for k in producto_keys)), None)
            
            unidad_medida_keys = ['u.m venta', 'um venta', 'unidad venta', 'u. medida', 'umedida', 'um'] 
            col_unidad_medida = next((c for c in df_pre_filtros.columns if any(k in str(c).lower() for k in unidad_medida_keys)), None)
            
            condicion_keys = ['condicion', 'codigocondicion', 'cond']
            col_condicion = next((c for c in df_pre_filtros.columns if any(k in str(c).lower() for k in condicion_keys)), None)
            
            clase_factura_keys = ['clase de factura', 'clasefactura', 'clase_factura', 'clase.factura', 'cl.f'] 
            col_clase_factura = next((c for c in df.columns if any(k in str(c).lower().replace(' ', '') for k in clase_factura_keys)), None)

            if not col_cliente or not col_factura or not col_monto:
                st.error("Error: Revise los encabezados de su archivo (Solicitante/Cliente, Factura/Asignacion y Precio/Monto).")
                st.stop()
                
            if not col_producto:
                 col_producto = 'Material_Dummy'
                 df_pre_filtros[col_producto] = ''
            
            if col_producto in df_pre_filtros.columns:
                df_pre_filtros[col_producto] = df_pre_filtros[col_producto].astype(str).apply(clean_leading_zeros)

            if selected_portfolio_cod != '--':
                 allowed_classes = ALLOWED_INVOICE_CLASSES.get(selected_portfolio_cod, [])
                 if allowed_classes:
                    if col_clase_factura is None:
                        st.warning(f"Advertencia: No se encontró la columna 'Clase de Factura' (Cl.F.) en el archivo. No se aplicará el filtro de Cl.F. para el Portafolio **{selected_portfolio_cod}**.")
                    elif col_clase_factura in df_pre_filtros.columns:
                        df_pre_filtros[col_clase_factura] = df_pre_filtros[col_clase_factura].astype(str).str.strip().str.upper()
                        original_rows = len(df_pre_filtros)
                        df_pre_filtros = df_pre_filtros[df_pre_filtros[col_clase_factura].isin(allowed_classes)].copy()
            
                        filtered_rows = len(df_pre_filtros)
                        
                        if filtered_rows == 0:
                            st.error(f"Error: No se encontraron facturas válidas en el archivo para el Portafolio **{selected_portfolio_cod}** con Clase de Factura (Cl.F.) **{', '.join(allowed_classes)}**.")
                            df_para_mostrar_editor = pd.DataFrame()
                            st.stop()
                        elif filtered_rows < original_rows:
                            pass 

            if col_factura:
                all_invoices_list = df_pre_filtros[col_factura].astype(str).fillna('').str.strip().unique().tolist()
                
                norm_to_originals = {}
                for orig in all_invoices_list:
                    digits = re.sub(r'\D', '', orig)
                    norm = clean_leading_zeros(digits)
                    if norm:
                        norm_to_originals.setdefault(norm, set()).add(orig)
                    else:
                        norm_to_originals.setdefault(orig.strip(), set()).add(orig)

                # --- LÓGICA DE SALDO RESTANTE Y ACTUALIZACIÓN VISUAL ---
                used_amount_map = {}
                if st.session_state['stacked_invoices']:
                    for batch_df in st.session_state['stacked_invoices']:
                        inv_col_name = 'ASIGNACION' if 'ASIGNACION' in batch_df.columns else col_factura
                        amount_col_name = 'Monto NC Asignado'

                        if inv_col_name in batch_df.columns and amount_col_name in batch_df.columns:
                            for idx, row in batch_df.iterrows():
                                inv_id = str(row[inv_col_name]).strip()
                                val = convert_value_to_float(row[amount_col_name])
                                if val:
                                    used_amount_map[inv_id] = used_amount_map.get(inv_id, 0.0) + val

                if used_amount_map:
                    df_pre_filtros['__monto_numeric__'] = df_pre_filtros[col_monto].apply(convert_value_to_float)
                    
                    def reduce_balance(group):
                        inv_id = str(group.name).strip() 
                        
                        if inv_id in used_amount_map:
                            total_used = used_amount_map[inv_id]
                            for idx in group.index:
                                if total_used <= 0.01: 
                                    break
                                
                                current_val = group.at[idx, '__monto_numeric__']
                                if pd.isna(current_val): continue
                                
                                if current_val > total_used:
                                    group.at[idx, '__monto_numeric__'] = current_val - total_used
                                    total_used = 0
                                else:
                                    group.at[idx, '__monto_numeric__'] = 0
                                    total_used -= current_val
                        return group

                    df_pre_filtros = df_pre_filtros.groupby(col_factura, group_keys=False).apply(reduce_balance)
                    
                    # Filtramos las que ya no tienen saldo
                    df_pre_filtros = df_pre_filtros[df_pre_filtros['__monto_numeric__'] > 0.01].copy()
                    
                    # --- AQUÍ ESTÁ EL CAMBIO IMPORTANTE: Actualizar la columna VISUAL ---
                    df_pre_filtros[col_monto] = df_pre_filtros['__monto_numeric__'].apply(format_monto_local)
                
                # ---------------------------------------------------------------------------------

                referenced_originals = set()
                for idx, row in df_pre_filtros.iterrows():
                    row_invoice = str(row.get(col_factura, '')).strip()
                    for col_name in df_pre_filtros.columns:
                        try:
                            cell_content = str(row[col_name]).strip()
                        except Exception:
                            cell_content = ''
                        if not cell_content:
                            continue

                        for num_in_cell in re.findall(r'\d{7,}', cell_content):
                            norm_num = clean_leading_zeros(num_in_cell)
                            if norm_num in norm_to_originals:
                                for original_invoice in norm_to_originals[norm_num]:
                                    if str(original_invoice).strip() != row_invoice:
                                        referenced_originals.add(original_invoice)

                if referenced_originals:
                    initial_rows = len(df_pre_filtros)
                    df_pre_filtros = df_pre_filtros[~df_pre_filtros[col_factura].astype(str).str.strip().isin(referenced_originals)].copy()
                    final_rows = len(df_pre_filtros)

            df_temp_validation = df_pre_filtros.copy()
            
            if client_code_list and col_cliente and col_cliente in df_temp_validation.columns:
                df_temp_validation[col_cliente] = df_temp_validation[col_cliente].astype(str)
                
                if not df_temp_validation[col_cliente].isin(client_code_list).any():
                    st.error(f"Error: Ninguno de los Códigos de Cliente ingresados ({', '.join(client_code_list)}) existe en el archivo cargado (o se filtró por Cl.F.).")
                    st.stop()
            
            if monto_nc is not None and monto_nc > 0:
                
                if not client_code_list:
                    st.error(" **ERROR:** Debe ingresar al menos un **Código de Cliente** para que el sistema pueda buscar las facturas que cubran el Monto de la Nota de Crédito.")
                    df_para_mostrar_editor = pd.DataFrame()
                    st.stop()
                
                df_temp_all_lines = df_pre_filtros.copy()
                
                if client_code_list and col_cliente and col_cliente in df_temp_all_lines.columns:
                    df_temp_all_lines[col_cliente] = df_temp_all_lines[col_cliente].astype(str)
        
                    df_temp_all_lines = df_temp_all_lines[df_temp_all_lines[col_cliente].isin(client_code_list)].copy()
                    
                    if df_temp_all_lines.empty:
                        st.error(f"No se encontraron facturas para los clientes: {', '.join(client_code_list)} (o fueron filtradas por Cl.F.).")
                        st.stop()
                
                invoices_with_product = []
        
                if product_code_list and col_producto and col_producto in df_temp_all_lines.columns:
                    
                    df_product_filtered = df_temp_all_lines[df_temp_all_lines[col_producto].astype(str).isin(product_code_list)].copy()
                    
                    if df_product_filtered.empty:
                        if product_code_list: 
                            st.error(f"No se encontraron facturas para los clientes **{', '.join(client_code_list)}** que también contengan alguno de los Productos **{', '.join(product_code_list)}**. No se puede realizar la asignación.")
                            df_para_mostrar_editor = pd.DataFrame()
                            st.stop()
                            
                    invoices_with_product = df_product_filtered[col_factura].unique().tolist()
                    
                    product_code_final_str = product_code_list[0] if product_code_list else ''

                    if product_code_list:
                        df_temp_for_coverage = df_temp_all_lines[df_temp_all_lines[col_factura].isin(invoices_with_product)].copy()
                    else:
                        df_temp_for_coverage = df_temp_all_lines.copy()
                else:
                    product_code_final_str = ''
                    df_temp_for_coverage = df_temp_all_lines.copy()

                df_temp_for_coverage['__monto_numeric__'] = df_temp_for_coverage[col_monto].apply(convert_value_to_float)
                df_temp_for_coverage.dropna(subset=['__monto_numeric__'], inplace=True)
                
                if df_temp_for_coverage.empty:
                     st.info("No hay facturas que coincidan con los filtros, Monto y Cl.F. de Portafolio.")
                     df_para_mostrar_editor = pd.DataFrame() 
                else:
                    with st.spinner("Calculando..."):
                        chosen_invoices_df, cliente_a_usar, monto_cubierto = find_invoices_by_total_sum(df_temp_for_coverage, monto_nc, col_factura, col_monto, col_cliente, col_producto, assignment_mode)
                    
                    if chosen_invoices_df is None:
                        monto_faltante = monto_nc - monto_cubierto
                        error_message = f"Error de Cobertura: El monto de NC de **{format_monto_local(monto_nc)}** no se pudo cubrir."
                        filtro_aplicado = f" (filtradas por clientes: {', '.join(client_code_list)}, Cl.F. y facturas con productos: {', '.join(product_code_list)})" if product_code_list else f" (filtradas por clientes: {', '.join(client_code_list)} y Cl.F.)"
                        error_message += f"\nEl monto total de **todas las facturas disponibles**{filtro_aplicado} es solo **{format_monto_local(monto_cubierto)}**."
                        error_message += f"\n**Monto Faltante:** {format_monto_local(monto_faltante)}"
                        st.error(error_message)
                        df_para_mostrar_editor = pd.DataFrame()
                        
                    elif not chosen_invoices_df.empty:
                        num_invoices = len(chosen_invoices_df)
                            
                        df_para_mostrar_editor = chosen_invoices_df.copy()
                        df_para_mostrar_editor['Monto Filas Selecc.'] = df_para_mostrar_editor['total_sum']
                        total_available_sum = df_para_mostrar_editor['Monto Filas Selecc.'].sum()
                        
                        if total_available_sum == 0:
                            st.error("Error: La suma de los montos de las facturas seleccionadas es cero. Revise su archivo.")
                            df_para_mostrar_editor = pd.DataFrame()
                        else:
                            if assignment_mode == 'Prorrateo (Recomendado)':
                                if len(df_para_mostrar_editor) == 1 and df_para_mostrar_editor['Monto Filas Selecc.'].iloc[0] >= monto_nc:
                                    df_para_mostrar_editor['Monto NC Asignado'] = monto_nc
                                else:
                                    prorate_factor = monto_nc / monto_cubierto if monto_cubierto != 0 else 0
                                    df_para_mostrar_editor['Monto NC Asignado'] = df_para_mostrar_editor['total_sum'] * prorate_factor
                                    df_para_mostrar_editor['Monto NC Asignado'] = df_para_mostrar_editor['Monto NC Asignado'].round(2)
            
                                    assigned_sum = df_para_mostrar_editor['Monto NC Asignado'].sum()
                                    difference = monto_nc - assigned_sum
                                    
                                    if not np.isclose(difference, 0):
                                        last_index = df_para_mostrar_editor.index[-1]
                                        df_para_mostrar_editor.loc[last_index, 'Monto NC Asignado'] += difference
                                        df_para_mostrar_editor.loc[last_index, 'Monto NC Asignado'] = df_para_mostrar_editor.loc[last_index, 'Monto NC Asignado'].round(2) 
                                        
                                    df_para_mostrar_editor = df_para_mostrar_editor.drop(columns=['total_sum'], errors='ignore')
                                    
                            elif assignment_mode == 'Estricto (Truncar)':
                                df_para_mostrar_editor = df_para_mostrar_editor.drop(columns=['total_sum'], errors='ignore')
                            
                            if product_code_final_str and col_producto not in df_para_mostrar_editor.columns:
                                df_para_mostrar_editor[col_producto] = product_code_final_str
                            elif col_producto in df_para_mostrar_editor.columns:
                                if product_code_final_str:
                                     df_para_mostrar_editor[col_producto] = df_para_mostrar_editor[col_producto].replace('', product_code_final_str).fillna(product_code_final_str)
                            else:
                                 df_para_mostrar_editor[col_producto] = ''

                            df_para_mostrar_editor[col_cliente] = cliente_a_usar 
                            
                            df_para_mostrar_editor['Peso %'] = df_para_mostrar_editor['Monto NC Asignado'] / monto_nc * 100.0 if monto_nc != 0 else 0.0
                            df_para_mostrar_editor['Cantidad'] = '1'
                            df_para_mostrar_editor['VARIACION DE PRECIO'] = df_para_mostrar_editor['Monto NC Asignado']
                            
                            if condicion_forzada:
                                df_para_mostrar_editor['CONDICION'] = condicion_forzada
                            else:
                                df_para_mostrar_editor['CONDICION'] = 'ZNOT'
                                
                            if col_unidad_medida and col_unidad_medida in df_temp_for_coverage.columns:
                                df_um_candidates = df_temp_for_coverage.copy()
                                if product_code_list: 
                                    df_um_candidates = df_temp_all_lines.copy()
                                    df_um_candidates = df_um_candidates[df_temp_all_lines[col_producto].astype(str).isin(product_code_list)]
                                    
                                if not df_um_candidates.empty:
                                    um_map = df_um_candidates.drop_duplicates(subset=[col_factura], keep='first').set_index(col_factura)[col_unidad_medida].to_dict()
                                    df_para_mostrar_editor['U. MEDIDA'] = df_para_mostrar_editor[col_factura].map(um_map).fillna('UN')
                                else:
                                    df_para_mostrar_editor['U. MEDIDA'] = 'UN'
                                    
                            elif 'U. MEDIDA' not in df_para_mostrar_editor.columns:
                                df_para_mostrar_editor['U. MEDIDA'] = 'UN'
                                
                            st.session_state['df_for_export_single_line'] = df_para_mostrar_editor.copy() 
                            
                        if st.session_state.get('df_for_export_single_line') is None or df_para_mostrar_editor.empty:
                             st.error("Error al construir la tabla de resultados después del cálculo.")
            else: 
                df_filtrado = df_pre_filtros.copy()
                
                if client_code_list and col_cliente and col_cliente in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado[col_cliente].isin(client_code_list)].copy()
                    
                if product_code_list and col_producto and col_producto in df_filtrado.columns:
                    df_filtrado = df_filtrado[df_filtrado[col_producto].astype(str).isin(product_code_list)].copy()

                if df_filtrado.empty:
                     st.info("No hay facturas que coincidan con los filtros (o fueron filtradas por Cl.F.).")
                     df_para_mostrar_editor = pd.DataFrame()
                else:
                    if col_factura and col_factura in df_filtrado.columns:
                        unique_invoices = df_filtrado[col_factura].unique()
                        df_para_mostrar = df_filtrado.drop_duplicates(subset=[col_factura]).copy()
                    else:
                        st.error("Columna de factura no encontrada para la visualización.")
                        df_para_mostrar_editor = pd.DataFrame()
                        st.stop()
                    
                    if condicion_forzada:
                        df_para_mostrar['CONDICION'] = condicion_forzada
                    elif col_condicion and col_condicion in df_para_mostrar.columns:
                        df_para_mostrar['CONDICION'] = df_para_mostrar[col_condicion].fillna(template_condicion)
                    elif 'CONDICION' not in df_para_mostrar.columns:
                        df_para_mostrar['CONDICION'] = template_condicion
                        
                    df_para_mostrar['Monto NC Asignado'] = ''
                    df_para_mostrar['VARIACION DE PRECIO'] = ''
                    
                    df_all_client_invoices = df_pre_filtros.copy()
                    
                    if client_code_list and col_cliente and col_cliente in df_all_client_invoices.columns:
                        df_all_client_invoices = df_all_client_invoices[df_pre_filtros[col_cliente].isin(client_code_list)].copy()
                    
                    df_all_client_invoices['__monto_numeric__'] = df_all_client_invoices[col_monto].apply(convert_value_to_float)
                    df_all_client_invoices.dropna(subset=['__monto_numeric__'], inplace=True)
                    
                    if not df_all_client_invoices.empty and col_factura in df_all_client_invoices.columns:
                        invoice_sums_dict = df_all_client_invoices.groupby(col_factura)['__monto_numeric__'].sum().to_dict()
                        df_para_mostrar['Monto Filas Selecc.'] = df_para_mostrar[col_factura].map(invoice_sums_dict)
                    else:
                        df_para_mostrar['Monto Filas Selecc.'] = 0 
                    
                    if 'Monto NC Asignado' in df_para_mostrar.columns:
                         df_para_mostrar = df_para_mostrar.drop(columns=['Monto NC Asignado', 'VARIACION DE PRECIO'], errors='ignore')
        
                    df_para_mostrar_editor = df_para_mostrar.copy()

        except Exception as e:
            st.error(f"Ocurrió un error: {e}")

        if not df_para_mostrar_editor.empty:
            df_para_mostrar_editor = df_para_mostrar_editor.rename(columns={
                col_factura: 'ASIGNACION',
                col_cliente: 'Solicitante',
                col_producto: 'Material'
            }, errors='ignore')
    
            if 'ASIGNACION' not in df_para_mostrar_editor.columns and col_factura in df.columns:
                 df_para_mostrar_editor['ASIGNACION'] = df_para_mostrar_editor[col_factura]
            if 'Solicitante' not in df_para_mostrar_editor.columns and col_cliente in df.columns:
                df_para_mostrar_editor['Solicitante'] = df_para_mostrar_editor[col_cliente]

            if cliente_a_usar:
                df_para_mostrar_editor['Solicitante'] = cliente_a_usar
            
            if monto_nc is None or monto_nc <= 0:
                if 'Material' in df_para_mostrar_editor.columns:
                     df_para_mostrar_editor['Material'] = df_para_mostrar_editor['Material'] 

            if 'Material' not in df_para_mostrar_editor.columns:
                df_para_mostrar_editor['Material'] = ''

            if col_unidad_medida and col_unidad_medida in df_para_mostrar_editor.columns:
                df_para_mostrar_editor['U. MEDIDA'] = df_para_mostrar_editor[col_unidad_medida].fillna('UN')
            elif 'U. MEDIDA' not in df_para_mostrar_editor.columns:
                df_para_mostrar_editor['U. MEDIDA'] = 'UN'
            if 'CONDICION' not in df_para_mostrar_editor.columns:
                df_para_mostrar_editor['CONDICION'] = 'ZNOT'
            if 'Observación' in df_para_mostrar_editor.columns:
                df_para_mostrar_editor = df_para_mostrar_editor.drop(columns=['Observación'], errors='ignore')
                
            df_para_mostrar_editor['Observación'] = ''
            df_para_mostrar_editor['Motivo'] = 'R02'
            
            if 'Canal' in df_para_mostrar_editor.columns and 'Canal de Distribucion' not in df_para_mostrar_editor.columns:
                 df_para_mostrar_editor = df_para_mostrar_editor.rename(columns={'Canal': 'Canal de Distribucion'}, errors='ignore')
                 
            current_date = datetime.now().date()
            date_to_use_str = current_date.strftime('%d/%m/%Y')
            date_cols_to_add = ["Fecha de Pedido", "Fecha de Precio", "Fecha de Factura"]
            
            for col in date_cols_to_add:
                if col not in df_para_mostrar_editor.columns:
                    df_para_mostrar_editor[col] = ''
                df_para_mostrar_editor[col] = date_to_use_str
                 
            selected_motivo = st.session_state.get('filtro_motivo')
            
            if selected_motivo == "Sin Motivo":
                texto_cabecera = "NCF.1"
            else:
                texto_cabecera = f"NCF.1 {selected_motivo}"
            
            if ticket_number:
                texto_cabecera += f" (Ticket {ticket_number})"
            df_para_mostrar_editor['TEXTO CABECERA'] = texto_cabecera
            
            df_para_mostrar_editor['Pedido Cliente'] = texto_cabecera 
            
            selected_defaults = PORTFOLIO_DEFAULTS.get(selected_portfolio_cod)

            if selected_defaults:
                df_para_mostrar_editor = df_para_mostrar_editor.assign(**selected_defaults)
                
            df_display_editor = df_para_mostrar_editor.copy()
            
            for col_name in ['Monto Filas Selecc.', 'Monto NC Asignado']:
                if col_name in df_display_editor.columns:
                    df_display_editor[col_name] = df_display_editor[col_name].apply(convert_value_to_float).apply(format_monto_local)
        
            if monto_nc is None or monto_nc <= 0:
                if 'Monto Filas Selecc.' in df_display_editor.columns:
                    df_display_editor['__sort_col__'] = df_display_editor['Monto Filas Selecc.'].apply(convert_value_to_float)
                    df_display_editor = df_display_editor.sort_values(by='__sort_col__', ascending=False).drop(columns=['__sort_col__'], errors='ignore')
                
            df_display_editor_renamed = df_display_editor.rename(columns={
                'Solicitante': 'Cod. Cliente',
                'ASIGNACION': 'Factura',
                'Material': 'Cod. Producto',
                'Monto NC Asignado': 'Monto Nota de Crédito',
                'U. MEDIDA': 'U. Medida',
                'Monto Filas Selecc.': 'Monto Total Factura' 
            }, errors='ignore')
            
            DISPLAY_COLS_FINAL = [
                'Cod. Cliente',
                'Factura',
                'Cod. Producto',
                'U. Medida',
                'Monto Total Factura', 
                'Monto Nota de Crédito'
            ]
            
            existing_cols = [col for col in DISPLAY_COLS_FINAL if col in df_display_editor_renamed.columns]
            df_display_editor_filtered = df_display_editor_renamed[existing_cols].copy()
            
            for col in DISPLAY_COLS_FINAL:
                if col not in df_display_editor_filtered.columns:
                    df_display_editor_filtered[col] = ''
                    
            st.dataframe(
                df_display_editor_filtered,
                column_config=col_config_dict, 
                hide_index=True,
                use_container_width=True,
            )
            
            df_export_base = df_para_mostrar_editor.copy()
            
            if not df_export_base.empty:
                df_for_all_export = df_export_base.drop(columns=['Monto Filas Selecc.', 'Peso %'], errors='ignore').copy()
                
                if st.button("Añadir Ticket", use_container_width=True):
                    ticket_for_stack_id = ticket_number if ticket_number else f"SINTICKET_{datetime.now().strftime('%H%M%S')}_{len(st.session_state['stacked_invoices'])}"
                    df_for_all_export['ID_Apilado'] = f"NC_{ticket_for_stack_id}"
                    st.session_state['stacked_invoices'].append(df_for_all_export.copy())
                    st.success("Ticket añadido.")
                    st.rerun() 
            else:
                st.info("No hay facturas en la tabla de resultados para la descarga.")

with tab2:
    if not st.session_state['stacked_invoices']:
        pass
    else:
        st.header("Tickets Generados")
        stacked_col_config_dict = {
            "ID de Lote": st.column_config.TextColumn("ID de Lote", help="Identificador único del ticket generado", width="medium"),
            "Cod. Cliente": st.column_config.TextColumn("Cod. Cliente", width="small", help="Código del Cliente (Solicitante)"),
            "Factura": st.column_config.TextColumn("Factura", width="small", help="Número de Factura o Documento de Asignación"),
            "Cod. Producto": st.column_config.TextColumn("Cod. Producto", width="small", help="Código del Material o Producto"),
            "U. Medida": st.column_config.TextColumn("U. Medida", width="small", help="Unidad de Medida (Ej: UN, KG)"),
            "Monto Nota de Crédito": st.column_config.TextColumn("Monto Nota de Crédito", width="small", help="Monto de la NC asignado (prorrateado) a esta factura."),
        }
        
        for i, df_lote in reversed(list(enumerate(st.session_state['stacked_invoices']))):
            lote_id = f"Lote Vacío {i}" 
            if not df_lote.empty:
                if 'TEXTO CABECERA' in df_lote.columns and not df_lote['TEXTO CABECERA'].empty:
                    header_text = str(df_lote['TEXTO CABECERA'].iloc[0]).strip()
                    if header_text:
                        lote_id = header_text
                elif 'ID_Apilado' in df_lote.columns and not df_lote['ID_Apilado'].empty:
                    lote_id = str(df_lote['ID_Apilado'].iloc[0]).strip()
                        
            with st.expander(f"{lote_id}"):
                df_lote_display = df_lote.copy()
                
                required_display_cols = ['Solicitante', 'ASIGNACION', 'Material', 'U. MEDIDA', 'Monto NC Asignado', 'ID_Apilado']
                for col in required_display_cols:
                    if col not in df_lote_display.columns:
                        df_lote_display[col] = ''
                        
                if 'Monto NC Asignado' in df_lote_display.columns:
                    df_lote_display['Monto NC Asignado'] = df_lote_display['Monto NC Asignado'].apply(convert_value_to_float).apply(format_monto_local)

                df_lote_display_renamed = df_lote_display.rename(columns={
                    'Solicitante': 'Cod. Cliente',
                    'ASIGNACION': 'Factura',
                    'Material': 'Cod. Producto',
                    'Monto NC Asignado': 'Monto Nota de Crédito',
                    'U. MEDIDA': 'U. Medida',
                    'ID_Apilado': 'ID de Lote'
                }, errors='ignore')

                stacked_display_cols_final = [
                    'ID de Lote', 
                    'Cod. Cliente',
                    'Factura',
                    'Cod. Producto',
                    'U. Medida',
                    'Monto Nota de Crédito'
                ]

                existing_stacked_cols = [col for col in stacked_display_cols_final if col in df_lote_display_renamed.columns]
                df_lote_display_filtered = df_lote_display_renamed[existing_stacked_cols].copy()

                st.dataframe(
                    df_lote_display_filtered,
                    column_config=stacked_col_config_dict,
                    hide_index=True,
                    use_container_width=True,
                )
                
                if st.button(f"Eliminar este Lote ({lote_id})", key=f"delete_lote_{i}", type="secondary"):
                    st.session_state['stacked_invoices'].pop(i)
                    st.success(f"Lote '{lote_id}' eliminado correctamente.")
                    st.rerun()

        col_clear_stack, col_download_stack = st.columns([1, 1])

        with col_download_stack:
            if st.session_state['stacked_invoices']:
                df_stacked_all_for_download = pd.concat(st.session_state['stacked_invoices'], ignore_index=True)
                selected_portfolio_for_download = st.session_state.get('portafolio_cod', '--')
                
                if not df_stacked_all_for_download.empty:
                    output_buffer_stacked = create_excel_for_all_invoices(df_stacked_all_for_download.drop(columns=['ID_Apilado'], errors='ignore'), selected_portfolio_for_download, multiple_invoices=True)
                else:
                    output_buffer_stacked = None
                
                if output_buffer_stacked:
                    st.download_button(
                        label="Descargar Todos los Tickets",
                        data=output_buffer_stacked,
                        file_name=get_file_name(selected_portfolio_for_download, "", multiple_invoices=True),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                else:
                    st.warning("No hay tickets para descargar o no se pudo generar el archivo.")

        with col_clear_stack:
            if st.button("Limpiar Todos los Tickets", use_container_width=True, type="secondary"):
                st.session_state['stacked_invoices'] = []
                st.success("Todos los tickets han sido limpiados.")
                st.rerun()
